import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from modules.colored_console import cprint
import os


def load_dict_key_row_from_excel(filename):  # Здесь ключ это шапка столбца
    try:
        df = pd.read_excel(filename)
        data = df.to_dict(orient='list')
        print(f"Данные успешно загружены из файла {filename}")
        return data
    except FileNotFoundError:
        cprint.error(f"Ошибка: Файл {filename} не найден.")
        return {}
    except Exception as e:
        cprint.error(f"Произошла ошибка при чтении файла {filename}: {e}")
        return {}


'''load_dict_key_column_from_excel 
- функция отличается от load_dict_key_row_from_excel: 
здесь ключи в столбце с заголовком index_of_column'''
def load_dict_key_column_from_excel(filename, index_of_column = 'Keys in Column'):  # Здесь ключи читаются из столбца с заголовком index_of_column
    try:
        # Чтение данных из файла Excel
        df = pd.read_excel(filename, index_col=index_of_column)
        # Преобразование DataFrame обратно в словарь
        data_dict = df.to_dict(orient='index')
        return data_dict
    except FileNotFoundError:
        cprint.error(f"ERROR!!! Файл {filename} не найден.")
    except Exception as e:
        cprint.error(f"ERROR!!! Произошла ошибка при чтении данных из файла: {e}")

'''Функция сохранения словаря в эксель. Ключи словаря - вначале строк, значения колонок переданы списком header'''
def save_dict_key_column_to_excel(data,  # словарь с данными
                                  filename,  # название файла с путем
                                  header,  # список с шапками колонок, кроме первой
                                  left_top_cell='Keys in Column',  # шапка первой колонки
                                  font_size=9,  # размер шрифта
                                  width_first_column=20,  # ширина первой колонки
                                  width_other_column=15,  # ширина остальных колонок
                                  width_last_column=10):  # ширина последней колонки

    try:
        df = pd.DataFrame.from_dict(data, orient='index', columns=header)
        df.index.name = left_top_cell

        if filename.endswith('.xlsx'):
            df.to_excel(filename, engine='openpyxl')
        elif filename.endswith('.xls'):
            df.to_excel(filename, engine='xlwt')
        else:
            raise ValueError("Unsupported file extension. Use '.xlsx' or '.xls'.")

        workbook = load_workbook(filename)
        sheet = workbook.active

        # Установка ширины столбцов
        sheet.column_dimensions['A'].width = width_first_column  # Ширина первого столбца
        for col in range(1, sheet.max_column):  # Ширина до предпоследнего столбца
            sheet.column_dimensions[chr(ord('A') + col)].width = width_other_column
        sheet.column_dimensions[chr(ord('A') + sheet.max_column - 1)].width = width_last_column  # Ширина последнего столбца

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.font = Font(size=font_size)
                # Создание объекта выравнивания
                alignment = Alignment(horizontal='center', vertical='center')
                # Применение выравнивания к ячейке
                cell.alignment = alignment

        workbook.save(filename)
        print(f"Данные успешно сохранены в файл {filename} с размером шрифта {font_size}")
    except PermissionError:
        cprint.error(f"Ошибка: Нет доступа к файлу {filename}. Возможно, он открыт в другой программе.")
    except Exception as e:
        cprint.error(f"Произошла ошибка при сохранении данных: {e}")

def save_dict_to_excel_without_header(data, filename, font_size=9):
    """
    Сохраняет словарь в файл Excel.

    :param data: Словарь, который нужно сохранить
    :param filename: Имя файла Excel
    :param font_size: Размер шрифта для ячеек
    """

    try:
        # Создаем DataFrame из словаря
        df = pd.DataFrame.from_dict(data, orient='index')

        # Определяем движок для сохранения файла в зависимости от расширения
        if filename.endswith('.xlsx'):
            df.to_excel(filename, header=False, engine='openpyxl')
        elif filename.endswith('.xls'):
            df.to_excel(filename, header=False, engine='xlwt')
        else:
            raise ValueError("Unsupported file extension. Use '.xlsx' or '.xls'.")

        # Открываем рабочую книгу для изменения шрифта
        workbook = load_workbook(filename)
        sheet = workbook.active

        # Применяем шрифт к каждой ячейке
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.font = Font(size=font_size)

        # Сохраняем изменения в файл
        workbook.save(filename)
        print(f"Данные успешно сохранены в файл {filename} с размером шрифта {font_size}")
    except PermissionError:
        cprint.error(f"Ошибка: Нет доступа к файлу {filename}. Возможно, он открыт в другой программе.")
    except Exception as e:
        cprint.error(f"Произошла ошибка при сохранении данных: {e}")


if __name__ == "__main__":
    # Пример словаря для сохранения
    data_to_save = {
        'Name': ['Alice', 'Bob', 'Charlie'],
        'Age': [25, 30, 35],
        'City': ['New York', 'Los Angeles', 'Chicago']
    }

    # Имя файла для сохранения данных
    filename = '../source/test_output_file_io.xlsx'

    # Сохранение словаря в файл Excel с изменением размера шрифта
    # save_dict_to_excel_without_header(data_to_save, filename, font_size=14)
    filename = '../source/dict_pair_in_exchange.xlsx'
    # Чтение данных из файла Excel
    loaded_data = load_dict_key_row_from_excel(filename)
    # print(loaded_data)

    print()
    filename = '../source/api_keys.xlsx'
    loaded_data = load_dict_key_column_from_excel(filename)
    print(loaded_data)
    # Вывод загруженных данных
    # print(loaded_data)
    print()
    # filename = '../source/api_keys.xlsx'
    # print(load_dict_key_column_from_excel(filename))
    # print(load_dict_key_column_from_excel(filename))
    print()
    # filename = '../source/test_output_file_io.xlsx'
    # print(save_dict_key_column_to_excel(data_to_save,filename, ['1','2','3']))

def save_dict_to_json_file(dictionary, filename, log = False):
    """
    Сохраняет словарь в файл JSON.

    :param dictionary: Словарь, который нужно сохранить.
    :param filename: Имя файла для сохранения.
    """
    try:
        with open(filename, 'w', encoding='utf-8') as file:
            json.dump(dictionary, file, ensure_ascii=False, indent=4)
        if log:
            cprint.success(f"Словарь успешно сохранен в файл {filename}")
    except Exception as e:
        cprint.error(f"Ошибка при сохранении словаря в файл: {e}")


import json


def load_dict_from_json_file(filename, log = False):
    """
    Восстанавливает словарь из файла JSON.

    :param filename: Имя файла для чтения.
    :return: Восстановленный словарь.
    """
    try:
        with open(filename, 'r', encoding='utf-8') as file:
            dictionary = json.load(file)
        if log:
            cprint.success(f"Словарь успешно восстановлен из файла {filename}")
        return dictionary

    except FileNotFoundError:
        cprint.error(f"ERROR!!! load_dict_from_json_file: Файл {filename} не найден.")
    except Exception as e:
        cprint.error(f"ERROR!!! load_dict_from_json_file при восстановлении словаря из файла: {e}")
        return None

def project_path():
    current_file_path = os.path.abspath(__file__)
    # Подымаемся на два уровня вверх (например, если ваш файл находится в 'project/subdir/script.py')
    project_root_path = os.path.abspath(os.path.join(current_file_path, '../../'))
    return project_root_path